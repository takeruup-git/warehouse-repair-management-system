import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def create_inspection_template():
    """
    定期自主検査記録表のExcelテンプレートを作成する
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "テンプレート"
    
    # スタイル定義
    header_font = Font(name='游ゴシック', bold=True, size=12)
    normal_font = Font(name='游ゴシック', size=10)
    title_font = Font(name='游ゴシック', bold=True, size=14)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # タイトル行
    ws.merge_cells('A1:J1')
    ws['A1'] = "フォークリフト定期自主検査記録表"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # ヘッダー行
    headers = [
        "管理番号", "メーカー", "機種", "機番", "製造年月日", "積載量", "揚高", "動力", "配置倉庫", "取扱担当者"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 検査項目セクション
    ws.merge_cells('A4:J4')
    ws['A4'] = "検査項目"
    ws['A4'].font = header_font
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A4'].fill = header_fill
    
    # 検査項目リスト
    inspection_items = [
        "1. 制動装置の異常の有無",
        "2. 操作装置の異常の有無",
        "3. 荷役装置の異常の有無",
        "4. 油圧装置の異常の有無",
        "5. 電気系統の異常の有無",
        "6. 車体の異常の有無",
        "7. 車輪の異常の有無",
        "8. 前照灯、方向指示器、警報装置等の異常の有無",
        "9. バッテリーの液量、比重の適否",
        "10. その他"
    ]
    
    for row, item in enumerate(inspection_items, start=5):
        cell = ws.cell(row=row, column=1)
        cell.value = item
        cell.font = normal_font
        cell.border = thin_border
        
        # 結果欄をマージ
        ws.merge_cells(f'B{row}:C{row}')
        result_cell = ws.cell(row=row, column=2)
        result_cell.border = thin_border
        
        # 備考欄をマージ
        ws.merge_cells(f'D{row}:J{row}')
        note_cell = ws.cell(row=row, column=4)
        note_cell.border = thin_border
    
    # 検査結果欄のヘッダー
    ws.cell(row=16, column=1).value = "検査実施日"
    ws.cell(row=16, column=1).font = header_font
    ws.cell(row=16, column=1).border = thin_border
    ws.cell(row=16, column=1).fill = header_fill
    
    ws.merge_cells('B16:D16')
    ws.cell(row=16, column=2).border = thin_border
    
    ws.cell(row=16, column=5).value = "検査実施者"
    ws.cell(row=16, column=5).font = header_font
    ws.cell(row=16, column=5).border = thin_border
    ws.cell(row=16, column=5).fill = header_fill
    
    ws.merge_cells('F16:J16')
    ws.cell(row=16, column=6).border = thin_border
    
    # 列幅の調整
    ws.column_dimensions['A'].width = 40
    for col in range(2, 11):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    # 行の高さ調整
    ws.row_dimensions[1].height = 30
    for row in range(2, 17):
        ws.row_dimensions[row].height = 20
    
    # テンプレートを保存
    template_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'templates')
    if not os.path.exists(template_dir):
        os.makedirs(template_dir)
    
    template_path = os.path.join(template_dir, 'forklift_inspection_template.xlsx')
    wb.save(template_path)
    
    return template_path

if __name__ == "__main__":
    create_inspection_template()