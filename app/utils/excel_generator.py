import os
import json
import openpyxl
from datetime import datetime
from flask import current_app
from app.utils.excel_template import create_inspection_template

def generate_form_from_uploaded_template(template_id, forklifts=None, blank=False):
    """
    アップロードされたExcelテンプレートを使用してフォームを生成する
    
    Args:
        template_id: FormTemplateのID
        forklifts: フォークリフトのリスト（Noneの場合は全台）
        blank: Trueの場合、空欄のフォームを生成
        
    Returns:
        生成されたExcelファイルのパス
    """
    from app.models.template import FormTemplate
    from app.models.forklift import Forklift
    
    # テンプレートを取得
    template = FormTemplate.query.get_or_404(template_id)
    template_path = os.path.join(current_app.root_path, template.file_path)
    
    # テンプレートが存在しない場合はエラー
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template file not found: {template_path}")
    
    # 出力ファイルのパスを設定
    output_dir = os.path.join(current_app.root_path, 'static', 'generated')
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(output_dir, f'{template.form_type}_form_{timestamp}.xlsx')
    
    # テンプレートを読み込む
    wb = openpyxl.load_workbook(template_path)
    
    # 空欄フォームの場合はそのまま保存
    if blank:
        wb.save(output_path)
        return os.path.join('static', 'generated', os.path.basename(output_path))
    
    # フォークリフトが指定されていない場合は全台取得
    if forklifts is None:
        forklifts = Forklift.query.filter_by(asset_status='active').all()
    
    # セルマッピングを取得
    cell_mapping = {}
    if template.cell_mapping:
        try:
            cell_mapping = json.loads(template.cell_mapping)
        except:
            pass
    
    # フォークリフトごとにシートを作成または更新
    for i, forklift in enumerate(forklifts):
        if i == 0:
            # 最初のフォークリフトは既存のシートを使用
            ws = wb.active
        else:
            # 2台目以降は新しいシートをコピーして作成
            source = wb.active
            target = wb.copy_worksheet(source)
            target.title = f"{forklift.management_number}"
            ws = target
        
        # フォークリフトデータをマッピングに従って埋め込む
        for field, cell in cell_mapping.items():
            if not cell:  # セルが指定されていない場合はスキップ
                continue
                
            value = None
            # ネストされた属性にアクセス（例: forklift.to_dict()['power_source_name']）
            if '.' in field:
                parts = field.split('.')
                obj = forklift
                for part in parts[:-1]:
                    if hasattr(obj, part):
                        obj = getattr(obj, part)
                    else:
                        break
                if hasattr(obj, parts[-1]):
                    value = getattr(obj, parts[-1])
            else:
                # 通常の属性アクセス
                if hasattr(forklift, field):
                    value = getattr(forklift, field)
                elif field in forklift.to_dict():
                    value = forklift.to_dict()[field]
            
            if value is not None:
                if isinstance(value, datetime.date):
                    value = value.strftime('%Y-%m-%d')
                elif isinstance(value, (int, float)) and field in ['load_capacity', 'lift_height']:
                    # 単位を追加
                    unit = 'kg' if field == 'load_capacity' else 'mm'
                    value = f"{value}{unit}"
                
                try:
                    ws[cell] = value
                except:
                    # セル参照が無効な場合はスキップ
                    pass
    
    # ファイルを保存
    wb.save(output_path)
    
    # 相対パスを返す（静的ファイルとしてアクセスするため）
    return os.path.join('static', 'generated', os.path.basename(output_path))

def generate_inspection_format(forklifts):
    """
    フォークリフト一覧から定期自主検査記録表のExcelファイルを生成する
    
    Args:
        forklifts: フォークリフトのリスト
        
    Returns:
        生成されたExcelファイルのパス
    """
    # テンプレートファイルのパスを取得
    template_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'templates')
    template_path = os.path.join(template_dir, 'forklift_inspection_template.xlsx')
    
    # テンプレートが存在しない場合は作成
    if not os.path.exists(template_path):
        template_path = create_inspection_template()
    
    # 出力ファイルのパスを設定
    output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'generated')
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(output_dir, f'forklift_inspection_{timestamp}.xlsx')
    
    # テンプレートを読み込む
    template_wb = openpyxl.load_workbook(template_path)
    template_ws = template_wb.active
    
    # 稼働中のフォークリフトのみを対象とする
    active_forklifts = [f for f in forklifts if f.asset_status == 'active']
    
    # テンプレートシートを削除
    if 'テンプレート' in template_wb.sheetnames:
        template_wb.remove(template_wb['テンプレート'])
    
    # フォークリフトごとにシートを作成
    for i, forklift in enumerate(active_forklifts):
        # シート名は管理番号を使用
        sheet_name = forklift.management_number
        # シート名の長さ制限（31文字）
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:28] + "..."
        
        # 既に同名のシートがある場合は連番を付ける
        if sheet_name in template_wb.sheetnames:
            j = 1
            while f"{sheet_name}_{j}" in template_wb.sheetnames and j < 100:
                j += 1
            sheet_name = f"{sheet_name}_{j}"
        
        # 新しいシートを作成
        if i == 0:
            ws = template_wb.create_sheet(sheet_name, 0)
        else:
            ws = template_wb.create_sheet(sheet_name)
        
        # テンプレートからスタイルをコピー
        for row in range(1, 17):
            for col in range(1, 11):
                src_cell = template_ws.cell(row=row, column=col)
                dst_cell = ws.cell(row=row, column=col)
                
                # セルの値とスタイルをコピー
                dst_cell.value = src_cell.value
                if src_cell.has_style:
                    dst_cell.font = src_cell.font
                    dst_cell.border = src_cell.border
                    dst_cell.fill = src_cell.fill
                    dst_cell.alignment = src_cell.alignment
        
        # マージセルをコピー
        for merged_cell_range in template_ws.merged_cells.ranges:
            ws.merge_cells(str(merged_cell_range))
        
        # フォークリフトデータを埋め込む
        ws.cell(row=3, column=1).value = forklift.management_number
        ws.cell(row=3, column=2).value = forklift.manufacturer
        ws.cell(row=3, column=3).value = forklift.model
        ws.cell(row=3, column=4).value = forklift.serial_number
        ws.cell(row=3, column=5).value = forklift.manufacture_date.strftime('%Y-%m-%d') if forklift.manufacture_date else ""
        ws.cell(row=3, column=6).value = f"{forklift.load_capacity}kg"
        ws.cell(row=3, column=7).value = f"{forklift.lift_height}mm"
        ws.cell(row=3, column=8).value = forklift.power_source_name
        ws.cell(row=3, column=9).value = f"{forklift.warehouse_group} {forklift.warehouse_number} {forklift.floor}"
        ws.cell(row=3, column=10).value = forklift.operator
        
        # 列幅の調整
        ws.column_dimensions['A'].width = 40
        for col in range(2, 11):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
        
        # 行の高さ調整
        ws.row_dimensions[1].height = 30
        for row in range(2, 17):
            ws.row_dimensions[row].height = 20
    
    # ファイルを保存
    template_wb.save(output_path)
    
    # 相対パスを返す（静的ファイルとしてアクセスするため）
    relative_path = os.path.join('static', 'generated', os.path.basename(output_path))
    return relative_path